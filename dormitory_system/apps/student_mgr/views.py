import os, csv, io
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View, TemplateView, FormView
from openpyxl import Workbook, load_workbook
from .models import Student, DormitoryRecord
from apps.dorm.models import Bed, Room, Building
from .forms import StudentForm


class AdminManagerMixin(UserPassesTestMixin):
    def test_func(self):
        user = self.request.user
        return user.is_authenticated and (user.is_super_admin() or user.is_manager())


# ========== 学生信息管理 ==========

class StudentListView(AdminManagerMixin, ListView):
    """学生列表"""
    model = Student
    template_name = 'student/list.html'
    context_object_name = 'students'
    paginate_by = 15

    def get_queryset(self):
        qs = Student.objects.all()
        kw = self.request.GET.get('keyword', '')
        gender = self.request.GET.get('gender', '')
        status = self.request.GET.get('status', '')
        if kw:
            qs = qs.filter(real_name__icontains=kw) | qs.filter(student_id__icontains=kw)
        if gender:
            qs = qs.filter(gender=gender)
        if status:
            qs = qs.filter(status=status)
        return qs


class StudentCreateView(AdminManagerMixin, CreateView):
    """添加学生"""
    model = Student
    form_class = StudentForm
    template_name = 'student/form.html'
    success_url = reverse_lazy('student_list')

    def form_valid(self, form):
        messages.success(self.request, f'学生 {form.instance.real_name} 添加成功')
        return super().form_valid(form)


class StudentUpdateView(AdminManagerMixin, UpdateView):
    """编辑学生"""
    model = Student
    form_class = StudentForm
    template_name = 'student/form.html'
    success_url = reverse_lazy('student_list')

    def form_valid(self, form):
        messages.success(self.request, '学生信息已更新')
        return super().form_valid(form)


class StudentDeleteView(AdminManagerMixin, DeleteView):
    """删除学生"""
    model = Student
    success_url = reverse_lazy('student_list')

    def get(self, *args, **kwargs):
        return self.delete(*args, **kwargs)


class StudentDetailView(AdminManagerMixin, DetailView):
    """学生详情"""
    model = Student
    template_name = 'student/detail.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['records'] = DormitoryRecord.objects.filter(student=self.object).order_by('-checkin_date')
        return ctx


# ========== 入住流程 ==========

class CheckInView(AdminManagerMixin, FormView):
    """学生入住"""
    template_name = 'student/checkin.html'

    def get_form(self, form_class=None):
        from django import forms
        class CheckInForm(forms.Form):
            building = forms.ModelChoiceField(
                queryset=Building.objects.filter(is_active=True),
                label='楼栋', widget=forms.Select(attrs={'class': 'form-control', 'id': 'id_building'})
            )
            room = forms.ModelChoiceField(
                queryset=Room.objects.none(),
                label='房间', widget=forms.Select(attrs={'class': 'form-control', 'id': 'id_room'})
            )
            bed = forms.ModelChoiceField(
                queryset=Bed.objects.none(),
                label='床位', widget=forms.Select(attrs={'class': 'form-control', 'id': 'id_bed'})
            )
            reason = forms.CharField(label='备注', required=False, widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 2}))
        return CheckInForm(**self.get_form_kwargs())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['student'] = get_object_or_404(Student, pk=self.kwargs['pk'])
        ctx['buildings'] = Building.objects.filter(is_active=True)
        return ctx

    def form_valid(self, form):
        student = get_object_or_404(Student, pk=self.kwargs['pk'])
        bed = form.cleaned_data['bed']

        if bed.status != '空闲':
            messages.error(self.request, '该床位已被占用')
            return self.form_invalid(form)

        if DormitoryRecord.objects.filter(student=student, status='入住中').exists():
            messages.error(self.request, '该学生已有在住记录，请先退宿')
            return self.form_invalid(form)

        # 创建入住记录
        DormitoryRecord.objects.create(
            student=student,
            bed=bed,
            status='入住中',
            operation_type='入住',
            operator=self.request.user,
            reason=form.cleaned_data.get('reason', ''),
        )
        bed.status = '已入住'
        bed.save()
        student.status = '在校'
        student.save()

        messages.success(self.request, f'{student.real_name} 入住成功')
        return redirect('student_detail', pk=student.pk)

    def get_success_url(self):
        return reverse_lazy('student_detail', kwargs={'pk': self.kwargs['pk']})


# ========== 调宿流程 ==========

class ChangeRoomView(AdminManagerMixin, FormView):
    """学生调宿"""
    template_name = 'student/change_room.html'

    def get_form(self, form_class=None):
        from django import forms
        class ChangeForm(forms.Form):
            building = forms.ModelChoiceField(
                queryset=Building.objects.filter(is_active=True),
                label='新楼栋', widget=forms.Select(attrs={'class': 'form-control', 'id': 'id_building'})
            )
            room = forms.ModelChoiceField(
                queryset=Room.objects.none(),
                label='新房间', widget=forms.Select(attrs={'class': 'form-control', 'id': 'id_room'})
            )
            bed = forms.ModelChoiceField(
                queryset=Bed.objects.none(),
                label='新床位', widget=forms.Select(attrs={'class': 'form-control', 'id': 'id_bed'})
            )
            reason = forms.CharField(label='调宿原因', required=True, widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 2}))
        return ChangeForm(**self.get_form_kwargs())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['student'] = get_object_or_404(Student, pk=self.kwargs['pk'])
        ctx['buildings'] = Building.objects.filter(is_active=True)
        return ctx

    def form_valid(self, form):
        student = get_object_or_404(Student, pk=self.kwargs['pk'])
        new_bed = form.cleaned_data['bed']
        old_record = DormitoryRecord.objects.filter(student=student, status='入住中').first()

        if not old_record:
            messages.error(self.request, '该学生当前没有在住记录')
            return self.form_invalid(form)

        if new_bed.status != '空闲':
            messages.error(self.request, '新床位已被占用')
            return self.form_invalid(form)

        old_bed = old_record.bed
        # 旧床位释放
        old_bed.status = '空闲'
        old_bed.save()
        old_record.status = '已调宿'
        old_record.save()

        # 新床位入住
        new_bed.status = '已入住'
        new_bed.save()
        DormitoryRecord.objects.create(
            student=student, bed=new_bed, status='入住中',
            operation_type='调宿', operator=self.request.user,
            reason=form.cleaned_data.get('reason', ''),
        )

        messages.success(self.request, f'{student.real_name} 调宿成功')
        return redirect('student_detail', pk=student.pk)


# ========== 退宿流程 ==========

class CheckOutView(AdminManagerMixin, View):
    """学生退宿"""
    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        record = DormitoryRecord.objects.filter(student=student, status='入住中').first()
        return render(request, 'student/checkout_confirm.html', {'student': student, 'record': record})

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        record = DormitoryRecord.objects.filter(student=student, status='入住中').first()
        if not record:
            messages.error(request, '该学生当前没有在住记录')
            return redirect('student_detail', pk=pk)

        from django.utils import timezone
        record.status = '已退宿'
        record.checkout_date = timezone.now()
        record.save()

        record.bed.status = '空闲'
        record.bed.save()

        student.status = '离校'
        student.save()

        messages.success(request, f'{student.real_name} 退宿成功')
        return redirect('student_detail', pk=pk)


# ========== 住宿记录 ==========

class DormRecordListView(AdminManagerMixin, ListView):
    model = DormitoryRecord
    template_name = 'student/dorm_records.html'
    context_object_name = 'records'
    paginate_by = 20

    def get_queryset(self):
        qs = DormitoryRecord.objects.select_related('student', 'bed__room__building')
        kw = self.request.GET.get('keyword', '')
        op = self.request.GET.get('operation', '')
        if kw:
            qs = qs.filter(student__real_name__icontains=kw) | qs.filter(student__student_id__icontains=kw)
        if op:
            qs = qs.filter(operation_type=op)
        return qs.order_by('-checkin_date')


class ChangeHistoryView(AdminManagerMixin, ListView):
    """调宿/退宿历史"""
    model = DormitoryRecord
    template_name = 'student/change_history.html'
    context_object_name = 'records'
    paginate_by = 20

    def get_queryset(self):
        return DormitoryRecord.objects.exclude(operation_type='入住').select_related(
            'student', 'bed__room__building'
        ).order_by('-checkin_date')


# ========== Excel导入导出 ==========

class ImportStudentsView(AdminManagerMixin, View):
    """Excel批量导入学生"""
    def get(self, request):
        return render(request, 'student/import.html')

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            messages.error(request, '请选择文件')
            return redirect('student_import')

        wb = load_workbook(file, read_only=True)
        ws = wb.active
        success = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                sid, name, gender, class_name, college, phone, parent_phone = row[:7]
                if not sid or not name:
                    continue
                Student.objects.get_or_create(
                    student_id=str(sid).strip(),
                    defaults={
                        'real_name': str(name).strip(),
                        'gender': str(gender).strip() if gender else '男',
                        'class_name': str(class_name).strip() if class_name else '',
                        'college': str(college).strip() if college else '',
                        'phone': str(phone).strip() if phone else '',
                        'parent_phone': str(parent_phone).strip() if parent_phone else '',
                    }
                )
                success += 1
            except Exception as e:
                errors.append(f'第{row[0]}行: {e}')

        msg = f'导入完成，成功 {success} 条'
        if errors:
            msg += f'，失败 {len(errors)} 条'
        messages.success(request, msg)
        return redirect('student_list')


class ExportStudentsView(AdminManagerMixin, View):
    """Excel导出学生"""
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = '学生信息'
        ws.append(['学号', '姓名', '性别', '班级', '学院', '手机号', '家长电话', '状态'])

        students = Student.objects.all()
        for s in students:
            ws.append([s.student_id, s.real_name, s.gender, s.class_name, s.college, s.phone, s.parent_phone, s.status])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)
        return response


class DownloadTemplateView(AdminManagerMixin, View):
    """下载导入模板"""
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = '导入模板'
        ws.append(['学号', '姓名', '性别', '班级', '学院', '手机号', '家长电话'])
        ws.append(['2024001', '张三', '男', '计算机一班', '计算机学院', '13800138000', '13900139000'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'
        wb.save(response)
        return response
